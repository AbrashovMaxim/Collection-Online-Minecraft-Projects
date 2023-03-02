from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Alignment, Font

def CreateWorkSheet(wb, curent_time):
    bys = f'{"0"+str(curent_time.day) if curent_time.day < 10 else curent_time.day}.{"0"+str(curent_time.month) if curent_time.month < 10 else curent_time.month}.{curent_time.year}'
    ws = wb.create_sheet(bys)

    cell = ws.cell(row=1, column=1)
    cell.value = 'Время'
    bold = Side(border_style="medium", color="000000")
    cell.border = Border(bold, bold, bold, bold)
    cell.fill = PatternFill("solid", fgColor="95B3D7")
    cell.font = Font(size=14, bold=bold)
    
    n = 30
    time_str = '00:00:00'
    date_format_str = '%H:%M:%S'
    given_time = datetime.strptime(time_str, date_format_str)
    first = True
    col = 2
    while True:
        if first == False and given_time.hour == 0 and given_time.minute == 0: break
        
        cell = ws.cell(row=1, column=col)
        a = f'{"0"+str(given_time.hour) if given_time.hour < 10 else given_time.hour}:{"0"+str(given_time.minute) if given_time.minute < 10 else given_time.minute}'
        cell.value=a
        cell.number_format = "HH:MM"
        cell.border = Border(bold, bold, bold, bold)
        cell.fill = PatternFill("solid", fgColor="95B3D7")
        cell.font = Font(size=14, bold=bold)
        cell.alignment = Alignment(horizontal='right')

        given_time = given_time + timedelta(minutes=n)
        col += 1
        first = False
    return wb

def SetStyleCells(cell, color, bold):
    thin = Side(border_style="thin", color="000000")

    if bold: cell.border = Border(bottom=thin, right=thin)
    else: cell.border = Border(right=thin, left=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(size=12, bold=bold)