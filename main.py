from datetime import datetime
import os
import time

import libs.excaliburcraft as excaliburcraft
import libs.cubixworld as cubixworld
import libs.mcskill as mcskill
import libs.pentacraft as pentacraft
import libs.loliland as loliland
import libs.simpleminecraft as simpleminecraft
import libs.gravitycraft as gravitycraft
import libs.shadowcraft as shadowcraft
import libs.gribland as gribland
import libs.minecraftonly as minecraftonly
import libs.sidemc as sidemc
import libs.grandgear as grandgear
import libs.mythicalworld as mythicalworld
import libs.borealis as borealis
import libs.victorycraft as victorycraft
import libs.centurymine as centurymine

from libs.pythonxl import *
from libs.logs import *

from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment

def SaveInFile(filename, data, curent_time):
    try: 
        wb = load_workbook('data/' + filename + '.xlsx')
        check_board = True
    except: 
        wb = CreateWorkSheet(Workbook(), curent_time)
        check_board = False
    try:
        if data == None: AppendLogs(curent_time, 'Ошибка'); return True
        if check_board:
            check = True
            for i in wb.worksheets:
                bys = f'{"0"+str(curent_time.day) if curent_time.day < 10 else curent_time.day}.{"0"+str(curent_time.month) if curent_time.month < 10 else curent_time.month}.{curent_time.year}'
                if i.title == bys:
                    if i['A2'].value == None:
                        row = 2;
                        for j,k in data.items():
                            comment = Comment(j, "Bot")
                            for l in k:
                                cell = i.cell(row=row, column=1)
                                cell.value=l[0]
                                cell.comment = comment
                                SetStyleCells(cell, "FABF8F", True)
                                row += 1
                        cell = i.cell(row=row, column=1)
                        cell.value = 'Общее кол-во игроков'
                        SetStyleCells(cell, "92CDDC", True)

                    bys = f'{"0"+str(curent_time.hour) if curent_time.hour < 10 else curent_time.hour}:{"0"+str(curent_time.minute) if curent_time.minute < 10 else curent_time.minute}'

                    row = i[1]
                    for j in row:
                        if str(j.value) == bys:
                            getColumn = j.column
                            row_data = 2
                            kolvo = 0
                            for k,l in data.items():
                                for m in l:
                                    cell = i.cell(row=row_data, column=getColumn)
                                    try:
                                        a = int(m[1])
                                        kolvo += a
                                        cell.value=a
                                    except:
                                        cell.value=m[1]
                                    cell.comment = Comment(m[2], "Bot")
                                    SetStyleCells(cell, "FCD5B4", False)
                                    row_data += 1
                            cell = i.cell(row=row_data, column=getColumn)
                            cell.value = kolvo
                            SetStyleCells(cell, "92CDDC", False)
                            break
                    check = False
                    break
            
            if check:
                wb = CreateWorkSheet(wb, curent_time)
                wb.save('data/'+ filename +'.xlsx')
                wb.close()
                SaveInFile(filename, data, curent_time)
            else:
                AppendLogs(curent_time, 'Save Successfully - [ '+ filename + ' ]')
                wb.save('data/'+ filename +'.xlsx')
                wb.close()
        else:
            try: wb.save('data/'+ filename +'.xlsx')
            except: 
                os.mkdir('data')
                wb.save('data/'+ filename +'.xlsx')
            wb.close()
            SaveInFile(filename, data, curent_time)
    except Exception as e:
        AppendLogs(curent_time, e)
    os.system('taskkill /f /im chrome.exe')

def main():
    cas = True
    CreateLogs()
    while True:
        curent_time = datetime.now()
        if (curent_time.minute == 30 or curent_time.minute == 0) and cas == True:
            AppendLogs(curent_time, 'Save Progress...')
            SaveInFile('ExcaliburCraft', excaliburcraft.GetExcaliburCraft(0, curent_time), curent_time)
            SaveInFile('CubixWorld', cubixworld.GetCubixWorld(0, curent_time), curent_time)
            SaveInFile('MCSkill', mcskill.GetMCSkill(0, curent_time), curent_time)
            SaveInFile('PentaCraft', pentacraft.GetPentaCraft(0, curent_time), curent_time)
            SaveInFile('LoliLand', loliland.GetLoliLand(0, curent_time), curent_time)
            SaveInFile('SimpleMinecraft', simpleminecraft.GetSimpleMinecraft(0, curent_time), curent_time)
            SaveInFile('GravityCraft', gravitycraft.GetGravityCraft(0, curent_time), curent_time)
            SaveInFile('ShadowCraft', shadowcraft.GetShadowCraft(0, curent_time), curent_time)
            SaveInFile('GribLand', gribland.GetGribLand(0, curent_time), curent_time)
            SaveInFile('MinecraftOnly', minecraftonly.GetMinecraftOnly(0, curent_time), curent_time)
            SaveInFile('SideMC', sidemc.GetSideMc(0, curent_time), curent_time)
            SaveInFile('GrandGear', grandgear.GetGrandGear(0, curent_time), curent_time)
            SaveInFile('MythicalWorld', mythicalworld.GetMythicalWorld(0, curent_time), curent_time)
            SaveInFile('Borealis', borealis.GetBorealis(0, curent_time), curent_time)
            SaveInFile('VictoryCraft', victorycraft.GetVictoryCraft(0, curent_time), curent_time)
            SaveInFile('CenturyMine', centurymine.GetCenturyMine(0, curent_time), curent_time)
            cas = False
        elif curent_time.minute != 30 or curent_time.minute != 0:
            cas = True
        AppendLogs(datetime.now(), 'Sleep')

        time.sleep(30)

if __name__ == "__main__":
    main()
    


# Type Server and Version: [[Server_Name_and_Version, kolvoPeople, kolvoSlots], [Server_Name_and_Version, kolvoPeople, kolvoSlots]]